from openpyxl import Workbook

#modal
from Modal_Student import getUPSubmittedStd,getStdUpOrder
from Modal_uv import getUv

class ExportUPData():
    def __init__(self,stream,UPDataSaveLocation):
        self.stream = stream
        self.UPDataSaveLocation = UPDataSaveLocation
        
        #
        self.save()
        
    #docHeader e.g[id,fullName,uvabrnames...,..]
    def setDocHeader(self,sheet):
        #get list of uv to get theri abreName for doc header
        uvLists = getUv(self.stream)
        
        #uv abre name staring column
        column = 3
        for uvList in uvLists:
            #uv abre name
            uvAbreName = uvList[2].upper()
            
            sheet.cell(row=1,column=1).value = 'Id'
            sheet.cell(row=1,column=2).value = 'Student Name'
            sheet.cell(row=1,column=column).value = uvAbreName
            column+=1
        
    #set data into each cells
    def setCellData(self,sheet):
        #get std list to get id,name from std
        stds = getUPSubmittedStd(self.stream)
        
        dataRow = 2
        for std in stds:
            #set id value
            sheet.cell(row=dataRow,column=1).value = std[0]
            #set name value
            sheet.cell(row=dataRow,column=2).value = std[1]
            #upOrder of std
            stdsUpOrder = getStdUpOrder(std[0])
            
            #upOrder starting column num
            column = 3
            for stdUpOrder in stdsUpOrder:
                #set upOrder for every uv
                sheet.cell(row=dataRow,column=column).value = stdUpOrder[0]
                column+=1

            dataRow+=1
    
    #save the doc into excel
    def save(self):
        wb = Workbook()
        sheet = wb.active
        
        self.setDocHeader(sheet)
        self.setCellData(sheet)
        
        wb.save(self.UPDataSaveLocation)
        
#ExportUPData('n','./upData-n.xlsx')
