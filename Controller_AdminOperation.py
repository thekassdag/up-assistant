#pyhton
from os import path
from datetime import datetime

#flask
from flask import request,url_for,redirect,flash

#library
from openpyxl import load_workbook

#modal
import Modal_Student

#module
from Module_exporetStdUPData import ExportUPData

def isStdListExcelTemp(excelFileName):
    wb = load_workbook(excelFileName,True)
    
    #the template must be 3col with 
    #Id,Student Full Name,Stream
    sheet = wb.active
    
    #totalRows and col
    rows = sheet.max_row
    cols = sheet.max_column
    
    if(cols == 3 and rows < 1):
        return False
    elif(cols == 3 and rows > 1):
        col1 = sheet['A1'].value.lower().lstrip()
        col2 = sheet['B1'].value.lower().lstrip()
        col3 = sheet['C1'].value.lower().lstrip()
        if(col1 == 'id' and col2 == 'student full name' and col3 == 'stream'):
            return True
        else:
            return False
    else:
        return False

def stdListExcel2Tuple(excelFileName):
    stdListTumble = []
    wb = load_workbook(excelFileName,True)
    sheet = wb.active

    #totalRows and col
    rows = sheet.max_row
    
    row = 2
    for row in range(2,rows):
        #id
        id = sheet.cell(row=row, column=1).value
        #name
        name = sheet.cell(row=row, column=2).value.upper()
        #stream
        stream = sheet.cell(row=row, column=3).value.upper()
        stdListTumble.append((name,id,stream,))
        row += 1
        
    return stdListTumble

def importStdList():
    acceptFileFormat = ('.csv','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet','application/vnd.ms-excel')

    stdListExcelFile = request.files['stdListData']
    
    #check if file is excel
    msgs = {}
    if stdListExcelFile.content_type not in acceptFileFormat:
        msgs = {'error':"import file format must be excel"}
    
    #if file is excel
    if(len(msgs) == 0):
        #rename the file
        file_name, file_extension = path.splitext(stdListExcelFile.filename)
        stdListExcelFile.filename = f'stdList{file_extension}'
        stdListExcelFile.save(path.join('./db',stdListExcelFile.filename))
        
        
        if(isStdListExcelTemp(path.join('./db',stdListExcelFile.filename)) == True):
            stdData = stdListExcel2Tuple(path.join('./db',stdListExcelFile.filename))
            Modal_Student.addStudents(stdData)
            msgs = {'success':'you have successfully imported student list into the system'}
        else:
            msgs = {'error':'you have imported system incompatible excel template'}
            
    for msg in msgs:
        flashMsg = msgs[msg]
        flash(flashMsg,msg)
    
    return redirect(url_for('adminDashboard'))

def exportUPData(stream):
    #used as id for the file
    fileNameExt = datetime.now().strftime('%S-%f')
    fileName = ''
    
    if(stream.upper() == 'N'):
        fileName = 'Natural-Student-University-Placement'
    else:
        fileName = 'Social-Student-University-Placement'
    
    #export it save to /download path
    exportedFile = f'{fileName}-{fileNameExt}.xlsx'
    ExportUPData(stream,f'../app/public/download/{exportedFile}')
    
    return exportedFile
